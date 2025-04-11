import datetime
import io
import json

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from app.collecting.models import TelegramMessage
from app.database import sync_session_maker
from fastapi.templating import Jinja2Templates

import pandas as pd
app = FastAPI()

@app.post("/webhook")
async def telegram_webhook(request: Request):
    data = await request.json()

    chat_id = None
    sender_first_name = None
    sender_last_name = None
    sender_username = None
    message_date = None
    topic_name = None

    if "message" in data:
        message = data["message"]
        chat = message.get("chat", {})
        chat_id = chat.get("id", None)

        sender = message.get("from", {})
        sender_first_name = sender.get("first_name", None)
        sender_last_name = sender.get("last_name", None)
        sender_username = sender.get("username", None)

        if "date" in message:
            message_date = datetime.datetime.fromtimestamp(message["date"])

        if "forum_topic_created" in message:
            forum_topic = message["forum_topic_created"]
            topic_name = forum_topic.get("name", None)

    try:
        telegram_message = TelegramMessage(
            chat_id=str(chat_id) if chat_id is not None else None,
            sender_first_name=sender_first_name,
            sender_last_name=sender_last_name,
            sender_username=sender_username,
            message_date=message_date,
            topic_name=topic_name,
            update_data=json.dumps(data)
        )
        sync_session_maker.add(telegram_message)
        sync_session_maker.commit()
    except Exception as e:
        sync_session_maker.rollback()
        print("Error saving message:", e)
    finally:
        sync_session_maker.close()

    return {"ok": True}



@app.get("/")
def redirect_to_chat():
    return RedirectResponse(url="/stats")

templates = Jinja2Templates(directory="templates")


@app.get("/stats", response_class=HTMLResponse)
async def stats_view(
        request: Request,
        start_date: str = None,
        end_date: str = None,
        topic_name: str = None
):
    session = sync_session_maker()

    topics_query = session.query(TelegramMessage.topic_name).distinct().all()
    topics = [t[0] for t in topics_query if t[0]]

    results = None
    if topic_name:
        query = session.query(
            TelegramMessage.sender_first_name,
            TelegramMessage.sender_last_name,
            TelegramMessage.sender_username,
            func.count(TelegramMessage.chat_id).label("message_count")
        ).filter(TelegramMessage.topic_name == topic_name)

        if start_date:
            try:
                start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                query = query.filter(TelegramMessage.message_date >= start_dt)
            except ValueError as ve:
                session.close()
                return HTMLResponse(content=f"Неверный формат начальной даты: {ve}", status_code=400)

        if end_date:
            try:
                end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d")
                end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
                query = query.filter(TelegramMessage.message_date <= end_dt)
            except ValueError as ve:
                session.close()
                return HTMLResponse(content=f"Неверный формат конечной даты: {ve}", status_code=400)

        query = query.group_by(
            TelegramMessage.sender_first_name,
            TelegramMessage.sender_last_name,
            TelegramMessage.sender_username,
        )
        results = query.all()

    session.close()

    return templates.TemplateResponse("stats.html", {
        "request": request,
        "topics": topics,
        "results": results,
        "start_date": start_date,
        "end_date": end_date,
        "selected_topic": topic_name
    })


@app.get("/stats/export")
async def export_stats(
        request: Request,
        start_date: str = None,  # Format: YYYY-MM-DD
        end_date: str = None,    # Format: YYYY-MM-DD
        topic_name: str = None
):
    session = sync_session_maker()

    if not topic_name:
        session.close()
        return HTMLResponse(content="No topic selected for export", status_code=400)

    query = session.query(
        TelegramMessage.sender_first_name,
        TelegramMessage.sender_last_name,
        TelegramMessage.sender_username,
        func.count(TelegramMessage.chat_id).label("message_count")
    ).filter(TelegramMessage.topic_name == topic_name)

    if start_date:
        try:
            start_dt = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(TelegramMessage.message_date >= start_dt)
        except ValueError as ve:
            session.close()
            return HTMLResponse(content=f"Invalid start date format: {ve}", status_code=400)

    if end_date:
        try:
            end_dt = datetime.datetime.strptime(end_date, "%Y-%m-%d")
            end_dt = end_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(TelegramMessage.message_date <= end_dt)
        except ValueError as ve:
            session.close()
            return HTMLResponse(content=f"Invalid end date format: {ve}", status_code=400)

    query = query.group_by(
        TelegramMessage.sender_first_name,
        TelegramMessage.sender_last_name,
        TelegramMessage.sender_username,
    )
    results = query.all()
    session.close()

    df = pd.DataFrame(
        results,
        columns=["First Name", "Last Name", "Username", "Message Count"]
    )

    time_info = ""
    time_part = ""
    if start_date and end_date:
        time_info = f"{start_date} to {end_date}"
        time_part = f"-{start_date}_{end_date}"
    elif start_date:
        time_info = f"from {start_date}"
        time_part = f"-{start_date}"
    elif end_date:
        time_info = f"until {end_date}"
        time_part = f"-{end_date}"

    file_name = f"Stats-{topic_name}{time_part}.xlsx"

    stream = io.BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=4, sheet_name="Statistics")

        workbook = writer.book
        worksheet = writer.sheets["Statistics"]

        last_col = worksheet.max_column
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "Statistics Report"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        worksheet.cell(row=2, column=1, value="Topic:")
        worksheet.cell(row=2, column=2, value=topic_name)
        worksheet.cell(row=3, column=1, value="Time Interval:")
        worksheet.cell(row=3, column=2, value=time_info)

        header_row = 5
        header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
        for col in range(1, last_col + 1):
            cell = worksheet.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        column_widths = {}
        for row in worksheet.iter_rows(min_row=header_row, max_row=worksheet.max_row, min_col=1, max_col=last_col):
            for cell in row:
                if cell.value:
                    column_widths[cell.column] = max(column_widths.get(cell.column, 0), len(str(cell.value)))
        for col, width in column_widths.items():
            worksheet.column_dimensions[get_column_letter(col)].width = width + 2

    stream.seek(0)
    headers = {
        'Content-Disposition': f'attachment; filename="{file_name}"'
    }

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )